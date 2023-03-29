# -*- coding: utf-8 -*-
# Copyright (C) 2014 Universidade de Aveiro, DETI/IEETA, Bioinformatics Group - http://bioinformatics.ua.pt/
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
import json
import logging
import re
from datetime import datetime

from Levenshtein import ratio
from django.db import transaction
from openpyxl import load_workbook

from fingerprint.models import Answer, AnswerChange
from qprocessors.choice import choice_list, serialize_list
from questionnaire.models import Choice, Question, QuestionSet, Questionnaire
from questionnaire.utils import QuestionNumber, TabularProcessor, next_free_slug
from searchengine.models import Slugs
from searchengine.search_indexes import convert_text_to_slug

logger = logging.getLogger(__name__)


class ImportQuestionnaire(object):
    """This class is used to import the fingerprint template
    """

    def __init__(self, file_path):
        self.file_path = file_path

    def import_questionnaire(self, preview, merge=None):
        raise NotImplementedError("Please Implement this method")

    def get_slug(self, slug, questionnaire):
        return next_free_slug(slug, create=False, scope=questionnaire)

    def format_number(self, number):
        # print number
        number_arr = number.split(".")

        result = number_arr[0] + "."
        for i in range(1,len(number_arr)):

            val = int(number_arr[i])
            if val<10:
                val = "0" + str(val)
            number_arr[i] = str(val)
            if (i!=len(number_arr)-1):
                result += str(val) + "."
            else:
                result += str(val)
        # print "result " + result
        return result

    """This method will build the object according with the type
    of the object to import.
    """
    @staticmethod
    def factory(t_type, file_path):
        if t_type == "excel":
            return ImportQuestionnaireExcel(file_path)

        else:
            raise Exception("The supplied format is not supported")

class CommentPlaceholder:
    value='comment'

class ImportQuestionnaireExcel(ImportQuestionnaire):

    QUESTION=0
    CATEGORY=1

    # choice match mode
    EXACT_MATCH=0
    SIMILARITY_MODE=1

    def __init__(self, file_path):
        ImportQuestionnaire.__init__(self, file_path)


    # this function implements the similarity algorithm, baed on a levenstein similarity algorithm
    # the idea is this algorithm be dynamically defined, for now its static
    def __isSimilar(self, comparing_option, options, percentage):

        closest = 0
        match = None

        for option in options:
            this_ratio = ratio(comparing_option,option)
            #print "'%s' and '%s' is %r similar" %(comparing_option, option, this_ratio)
            if this_ratio > closest and this_ratio > percentage:
                closest = this_ratio
                match = option

        return (closest, match)

    def __handleAnswerChanges(self, question, change_map, debug=False):
        if len(change_map) > 0:
            #print "ANSWERS:"

            def __answerChange(data, change_map):
                if data == None:
                    return serialize_list([])
                response = choice_list(data).values()
                for res in response:
                    try:
                        res['key'] = change_map[res['key']]
                    except KeyError:
                        pass

                return serialize_list(response)

            # Handle answer changes modifications
            answers = Answer.objects.filter(question=question)
            for ans in answers:
                ans.data = __answerChange(ans.data, change_map)

                if not debug:
                    ans.save()

                # Update answer history
                # For old values
                ans_history = AnswerChange.objects.filter(answer=ans)
                for hist in ans_history:
                    hist.old_value = __answerChange(hist.old_value, change_map)
                    hist.new_value = __answerChange(hist.new_value, change_map)

                    if not debug:
                        hist.save()

            # Update dependencies too
            '''print "UPDATE DEPENDENCIES FOR: %s" %(question.number)
            for key, value in change_map:
                print key
                print value
                print "--"
                dependent_questions = Question.objects.filter(checks__contains='dependent="%s,%s"' % (question.number, key))
                print "QUESTIONS DEPENDENT"
                print dependent_questions
            '''
            #raise Exception("CRAZY")

    def __processChoices(self, row, question, list_aux, mode=EXACT_MATCH, match_percentage=0.75, debug=False, infer_function=None):
        #_choices_array_aux=[]
        i = 0

        # get current questions if any
        old_choices = list(Choice.objects.filter(question=question).values_list('value', flat=True))

        #print old_choices
        change_map={}       # maps the changes to be made over the question
        indexes_map = {}    # since ordering is absolute, on first passthrough i generate the sortid numbers
        look_similar = []   # potential matches for 2nd pass similarity lookup
        maybe_new = []      # for 3rd pass we may infer

        # 1st pass: we do a first pass through to remove exact matches (is the same on both approaches)
        for ch in list_aux:
            has_opt = False
            clean_ch = ch

            if '{...}' in ch:
                clean_ch = ch.replace('{...}', '')
                has_opt = True

            i+=1
            indexes_map[clean_ch] = i

            if clean_ch in old_choices:
                choice = Choice.objects.get(question=question, value=clean_ch)
                choice.sortid = i
                choice.opt = has_opt

                if not debug:
                    choice.save()

                #_choices_array_aux.append(ch)
                old_choices.remove(clean_ch)

            else:
                look_similar.append(ch)


        def __similarMap(question, similar, ch):
            has_opt = False

            if '{...}' in ch:
                ch = ch.replace('{...}', '')
                has_opt = True

            change_map[similar] = ch

            choice=Choice.objects.get(question=question, value=similar)
            choice.text_en = ch
            choice.value = ch
            choice.sortid=indexes_map[clean_ch]
            choice.opt = has_opt


            if not debug:
                choice.save()

            #_choices_array_aux.append(ch)
            old_choices.remove(similar)

        # 2nd pass: lets analyse the rest that are not exact matches
        for ch in look_similar:
            has_opt = False
            clean_ch = ch

            if '{...}' in ch:
                clean_ch = ch.replace('{...}', '')
                has_opt = True

            if mode==self.SIMILARITY_MODE:

                (closest, similar) = self.__isSimilar(clean_ch, old_choices, match_percentage)
                # if considered similar
                if similar != None:
                    if closest < 1:
                        print "Replacing '%r' which is %r similar to '%r' on question %r" % (similar, closest, ch, question.number)

                    __similarMap(question, similar, ch)

                else:
                    # maybe be new, to check on 3rd pass
                    maybe_new.append(ch)

            # if this is exact match mode, we skip this step
            else:
                maybe_new.append(ch)

        # 3rd pass: if there's an boolean lambda infer function to non obvious cases dealing, run it
        run = list(maybe_new)
        if infer_function != None and len(maybe_new) > 0 and len(old_choices) > 0:
            for new in run:
                has_opt = False
                clean_new = new

                if '{...}' in new:
                    clean_new = new.replace('{...}', '')
                    has_opt = True

                print "RUN for " + str(new)
                if old_choices > 0:
                    for old in old_choices:

                        if infer_function(question.number, clean_new, old) == True:
                            print "Replacing '%r' which is user indicated similar to '%r' on question %r" % (old, new, question.number)
                            __similarMap(question, old, new)

                            maybe_new.remove(new)

                            #if we find a hit its done
                            break
                else:
                    print "No more old choices, others must be new"

        for ch in maybe_new:
            has_opt = False
            clean_ch = ch

            if '{...}' in ch:
                clean_ch = ch.replace('{...}', '')
                has_opt = True

            # otherwise we create a new entry
            print "Create new '%s'" %(clean_ch)
            try:
                choice = Choice(question=question, sortid=indexes_map[clean_ch], text_en=clean_ch, value=clean_ch, opt=has_opt)
                logger.debug("%s - Choice created %s ", row, choice)
                if not debug:
                    choice.save()
                #_choices_array_aux.append(ch)

                logger.debug("%s - Choice saved %s ", row, choice)
            except:
                logger.warning("%s - Error to save Choice %s", row, choice)
                raise

        if len(old_choices)> 0:
            print "REMOVED:"
            print old_choices
        # at last, we must remove the choices that dont appear in the new listing (considered removed)
        Choice.objects.filter(question=question, value__in=old_choices).delete()

        if mode==self.SIMILARITY_MODE:
            self.__handleAnswerChanges(question, change_map, debug=debug)


        return list_aux #_choices_array_aux

    def __processDisposition(self, disposition):
        if disposition == 'horizontal':
            return 1
        elif disposition == 'dropdown':
            return 2

        return 0

    def __handleQuestionNumber(self, level, qNumber, questionset):
        questionNumber = None

        if level.startswith('h'):
            questionNumber = qNumber.getNumber(level)
            questionNumber = self.format_number(str(questionNumber))

        else:
            questionNumber = level

            pos = level.split('.')
            poslen = len(pos)

            for question in questionset.questions():
                this_q = question.number.split('.')

                if poslen == len(this_q):

                    if pos[poslen-1] <= this_q[poslen-1]:
                        n = int(this_q[poslen-1])+1
                        if n < 10:
                            this_q[poslen-1] = '0'+str(n)
                        else:
                            this_q[poslen-1] = str(n)

                        question.number = ".".join(this_q)
                        question.save()

            #raise Exception('STOP THERE')

        return questionNumber


    def __getChoices(self, question):
        ''' Gets the choices_array from a question back into an choices_array.
            Useful on merge operation that point dependencies to questions already on the database
        '''
        if question.type in ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform']:
            choices = Choice.objects.get(question=question).values_list('value', flat=true)

            return choices

        elif question.type in ['choice-yesno', 'choice-yesnodontknow']:
            return ['yes', 'no', 'dontknow']

        return []


    def __processTabular(self, question, values):
        if values == None:
            values = ''
        tp = TabularProcessor(question)

        tp.append(values)

        return tp.json()

    def __handleQuestion(self, type, row,type_Column, level_number_column, text_question_Column, _questions_rows,
        _choices_array, qNumber, questionset, _checks, _debug, questionnaire, old_qsets_ids, mode=EXACT_MATCH, percentage=0.75, infer_function=None):

        question = None

        try:
            slug = None
            text_en = None
            if level_number_column.value.startswith('h'):
                text_en = str(level_number_column.value) + '. ' + str(text_question_Column.value)
            else:
                level = len(level_number_column.value.split('.'))-1

                text_en = 'h%s. %s' % (str(level),str(text_question_Column.value))


            dataType_column = None
            if type == self.CATEGORY:
                dataType_column = CommentPlaceholder()
            else:
                dataType_column = row[3]

            if row[7].value:
                slug = row[7].value
            else:
                slug = convert_text_to_slug(str(row[1].value)[:50])

                slug = self.get_slug(slug, questionnaire)

            slug = re.sub(r"\W", "_", slug)  # discard non-alphanumeric characters

            if row[5].value:
                helpText = row[5].value
            else:
                helpText = ''

            _tooltip = False

            if row[6].value:
                if str(row[6].value).lower() == 'yes':
                    _tooltip = True

            _advanced = True
            if len(row) > 12 and row[12].value:
                if str(row[12].value).lower() == 'no':
                    _advanced = False

            #If has dependencies
            if row[8].value:
                try:
                    dependencies_list = row[8]
                    list_dep_aux = dependencies_list.value.split('|')

                    list_dep_aux[0] = re.sub(r"\W", "_", list_dep_aux[0])  # discard non-alphanumeric characters

                    question_num_parent = None
                    try:
                        question_num_parent = _questions_rows.get(list_dep_aux[0]).number
                    except AttributeError:
                        ''' If this is a merge, the dependant question can already be on the questionset,
                            lets try looking for it
                        '''
                        try:
                            question = Question.objects.get(slug_fk__slug1=list_dep_aux[0],
                                questionset=questionset)
                            _questions_rows[list_dep_aux[0]] = question_num_parent

                            question_num_parent = question.number

                            _choices_array[list_dep_aux[0]] = self.__getChoices(question)

                        except Question.DoesNotExist:
                            raise Exception('The dependant with slug %s does not exist.' %(list_dep_aux[0]))




                    try:
                        index_aux = int(str(list_dep_aux[1]))-1
                        choice_parent_list = _choices_array.get(list_dep_aux[0])
                        choice_parent = choice_parent_list[index_aux]
                        _checks = 'dependent=\"%s,%s\"' % (str(question_num_parent), str(choice_parent))
                    except:
                        import sys, traceback
                        traceback.print_exc(file=sys.stderr)
                except:
                    raise

            try:
                questionNumber = self.__handleQuestionNumber(level_number_column.value, qNumber, questionset)

            except:
                if type==self.QUESTION:
                    logger.warning("%s - Error to create question number %s", type_Column.row, text_en)
                elif type==self.CATEGORY:
                    logger.warning("%s - Error to create Category number %s", type_Column.row, text_en)

                raise

            #print slug
            #Create or load slug
            slugs = Slugs.objects.filter(slug1=slug, description=text_en)
            if len(slugs) <= 0:
                slug_db = Slugs(slug1=slug, description=text_en)
                slug_db.save()
            else:
                slug_db = slugs[0]

            visible_default = False
            if row[10].value:
                if str(row[10].value).lower() == 'visible':
                    visible_default = True

            is_category=None
            is_stats=None

            if type==self.QUESTION:
                is_stats=True
                is_category=False
            elif type==self.CATEGORY:
                is_stats=False
                is_category=True

            try:
                question = Question.objects.get(slug_fk__slug1=slug_db.slug1, questionset__in=old_qsets_ids)

                question.questionset=questionset
                question.text_en=text_en
                question.number=str(questionNumber)
                question.type=dataType_column.value
                question.help_text=helpText
                question.stats=is_stats
                question.category=is_category
                question.tooltip=_tooltip
                question.checks=_checks
                question.visible_default=visible_default
                question.show_advanced=_advanced

            except Question.DoesNotExist:
                question = Question(questionset=questionset, text_en=text_en, number=str(questionNumber),
                                type=dataType_column.value, help_text=helpText, slug=slug, slug_fk=slug_db, stats=is_stats,
                                category=is_category, tooltip=_tooltip,
                                checks=_checks, visible_default=visible_default,
                                disposition=self.__processDisposition(row[11].value.lower()), show_advanced=_advanced)


            if dataType_column.value in ['open-validated']:
                ardict = {}

                if row[4].value:
                    # some basic types dont need regex
                    known_validations = {
                        "integer": "[+-]?\d+",
                        "decimal": "[+-]?\d*([.]\d*)?",
                        "scientific": "[+-]?\d*([.]\d*)?e[+-]?\d*([.]\d*)?",
                        "range": "[+\-]?\d*([.]\d*);[+\-]?\d*([.]\d*)",
                        "date": "\d{2}/\d{2}/\d{4}",
                        "time": "\d{2}:\d{2}:\d{2}",
                        "datetime": "\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}",
                        "text": ".*"
                    }
                    try:
                       ardict['regex'] = known_validations[row[4].value]
                       ardict['base'] = row[4].value
                    except KeyError:
                        # If this is not known, try to validate it as a regex
                        try:
                            re.compile(row[4].value)
                            ardict['regex'] = row[4].value
                        except re.error:
                            raise Exception("--ERROR: The regex on row %d, column 4 is not valid" % (type_Column.row))

                if row[5].value:
                    split = row[5].value.split('|')
                    lensplit = len(split)

                    if lensplit == 1:
                        ardict['unit'] = split[0]
                        question.help_text=""

                    elif lensplit == 2:
                        ardict['unit'] = split[0]
                        ardict['unit_desc'] = split[1]
                        question.help_text=""

                    elif lensplit == 3:
                        ardict['unit'] = split[0]
                        ardict['unit_desc'] = split[1]
                        question.help_text = split[2]

                    else:
                        raise Exception("-- ERROR: Invalid number of segments on help text row %d, column 5. Max syntax is unit|desc|help_text" % (type_Column.row))


                question.metadata = json.dumps(ardict)

            if not _debug:
                question.save()

                if type==self.QUESTION:
                    logger.debug("%s - Question created %s ", type_Column.row, question)
                elif type==self.CATEGORY:
                    logger.debug("%s - Category created %s ", type_Column.row, question)



            _questions_rows[slug] = question

            if type == self.QUESTION:
                if dataType_column.value in ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform']:
                    _choices_array_aux = []
                    # Parse of values list
                    values_list = row[4]
                    if (values_list!=None and values_list.value!=None):
                        list_aux = values_list.value.split('|')

                        _choices_array[slug] = self.__processChoices(type_Column.row, question, list_aux, debug=_debug,
                            mode=mode, match_percentage=percentage, infer_function=infer_function)

                elif dataType_column.value in ['choice-tabular']:
                    question.metadata = json.dumps({
                            'data': self.__processTabular(question, row[4].value)
                        })
                    if not _debug:
                        question.save()
                elif dataType_column.value in ['open-multiple-composition']:
                    tmpCompositionDefinition = row[4].value
                    question.metadata = tmpCompositionDefinition
                    if not _debug:
                        question.save()
                if dataType_column.value in ['choice-yesno',
                                                     'choice-yesnodontknow']:
                    _choices_array[slug] = ['yes', 'no', 'dontknow']

        except:
            logger.warning("%s - Error to save question %s", type_Column.row, text_en)

            raise

        return question

    @transaction.atomic
    def import_questionnaire(self, preview, merge=None, mode=EXACT_MATCH, percentage=0.75, infer_function=None):
        _debug = False

        included_questions = []

        qNumber = QuestionNumber()
        slugs = []

        wb = load_workbook(filename = self.file_path, data_only=True)
        ws = wb.get_active_sheet()

        old_qsets = []
        old_qsets_ids = []

        # Cell B1: Name of questionnaire
        name = ws.cell(row=1, column=2).value
        slugQ = convert_text_to_slug(name)
        disable = False

        questionnaire = None
        if merge != None:
            try:
                questionnaire = Questionnaire.objects.get(id=merge)

                old_qsets = list(questionnaire.questionsets())
                old_qsets_ids = list(questionnaire.questionsets().values_list('id', flat=True))

                questionnaire.in_preview = preview
                if not _debug:
                    questionnaire.qsets.clear()
                    questionnaire.save()
                    logger.debug("Questionnaire saved %s ", questionnaire)

            except Questionnaire.DoesNotExist:
                raise Exception('Questionnaire does not exist, so cant merge against it.')

        else:
            questionnaire = Questionnaire(name=name, disable=disable, slug=slugQ, redirect_url='/', in_preview=preview)
            logger.debug("Questionnaire created %s ", questionnaire)
            if not _debug:
                questionnaire.save()
                logger.debug("Questionnaire saved %s ", questionnaire)

        try:
            _choices_array = {}
            _questions_rows = {}

            #############################
            # TIPS:
            # Type of Row: QuestionSet, Category, Question
            # Columns: Type, Text/Question, Level/Number, Data type, Value list, Help text/Description, Tooltip, Slug, Stats
            #############################
            for row in ws.iter_rows(min_row=3):
                if len(row) > 0 and row[0].value != None:
                    type_Column = row[0]

                    text_question_Column = row[1]

                    if (text_question_Column.value!=None):
                        text_question_Column.value = text_question_Column.value.encode('ascii', 'ignore')
                    level_number_column = row[2]
                    _checks = ''

                    # Type = QUESTIONSET
                    # Columns required:  Type, Text/Question
                    # Columns optional:  Help text/Description, Tooltip
                    if str(type_Column.value) == "QuestionSet":

                        sortid = str(level_number_column.value)
                        try:
                            qNumber.getNumber('h0', sortid)
                        except:
                            raise
                        text_en = 'h1. %s' % text_question_Column.value

                        slug_qs = None
                        if row[7].value:
                            slug_qs = row[7].value
                        else:
                            slug_qs = str(slugQ) + "_" + convert_text_to_slug(str(text_question_Column.value))

                        if row[5].value:
                            helpText = row[5].value
                        else:
                            helpText = ""
                        tooltip = False
                        if row[6].value:
                            if str(row[6].value).lower() == 'yes':
                                tooltip = True

                        _advanced = True
                        try:
                            if row[12].value:
                                if str(row[12].value).lower() == 'no':
                                    _advanced = False
                        except:
                            _advanced = True

                        questionset = None
                        created = False

                        for qset in old_qsets:
                            if qset.heading == slug_qs:
                                questionset = qset
                                break

                        if questionset == None:
                            questionset = QuestionSet(questionnaire=questionnaire, sortid=sortid, heading=slug_qs,
                                checks='required', text_en=text_en, help_text=helpText, tooltip=tooltip, show_advanced=_advanced)
                            created=True

                        if created:
                            logger.debug("%s - QuestionSet created %s - %s ", type_Column.row, sortid, text_en)
                        else:
                            logger.debug("%s - QuestionSet retrieved %s - %s ", type_Column.row, sortid, text_en)

                            questionset.sortid=sortid
                            questionset.text_en=text_en
                            help_text=helpText
                            tooltip=tooltip
                            questionset.show_advanced=_advanced
                        try:
                            if not _debug:
                                questionset.save()

                                questionnaire.qsets.add(questionset)

                                logger.debug("%s - QuestionSet saved %s - %s ", type_Column.row, sortid, text_en)
                        except:

                            logger.warning("%s - Error to save questionset %s - %s", type_Column.row, sortid, text_en)

                            raise

                        #if not created:
                        #    last_question = Question.objects.filter(questionset=questionset).order_by('-id')[0]
                        #    qNumber.setState(last_question.number)

                    # Type = CATEGORY
                    # Columns required:  Type, Text/Question, Level/Number, Category
                    # Columns optional:  Help text/Description, Slug, Tooltip, Dependencies
                    elif str(type_Column.value) == "Category":
                        qstion = self.__handleQuestion(self.CATEGORY, row, type_Column, level_number_column, text_question_Column,
                            _questions_rows, _choices_array, qNumber, questionset, _checks, _debug,
                            questionnaire, old_qsets_ids, mode=mode, percentage=percentage, infer_function=infer_function)
                        try:
                            slugs = self.verify_slug(qstion.slug_fk.slug1,slugs,questionnaire)
                        except ValueError:
                            print("Slug: " + qstion.slug_fk.slug1)
                            raise ValueError("Not Unique Slug")
                        included_questions.append(qstion.id)
                    # Type = QUESTION
                    # Columns required:  Type, Text/Question, Level/Number, Data Type, Category, Stats
                    # Columns optional:  Value List, Help text/Description, Tooltip, Dependencies
                    else:
                        qstion = self.__handleQuestion(self.QUESTION, row, type_Column, level_number_column, text_question_Column,
                            _questions_rows, _choices_array, qNumber, questionset, _checks, _debug,
                            questionnaire, old_qsets_ids, mode=mode, percentage=percentage, infer_function=infer_function)
                        try:
                            slugs = self.verify_slug(qstion.slug_fk.slug1,slugs,questionnaire)
                        except ValueError:
                            print("Slug: " + qstion.slug_fk.slug1)
                            raise ValueError("Not Unique Slug")
                        included_questions.append(qstion.id)
        except:
            logger.warning("Error to save questionsets and questions of the questionnaire %s ", questionnaire)

            raise

        print "INCLUDED QUESTIONS LIST: "
        print included_questions
        print len(included_questions)
        print "-----"

        not_included = Question.objects.filter(questionset__in=old_qsets_ids).exclude(id__in=included_questions)

        print not_included.count()
        print "------"
        print "Removing not included questions (were discontinued):"
        for qst in not_included:
            print qst
            qst.delete()

        logger.debug("Questionnaire %s, questionsets, questions and choices created with success!! ", questionnaire)


        #raise Exception('Dont commit me dude')
        return questionnaire

    # Check if the slug is already used in the questionnaire
    def verify_slug(self,slug,slugs,questionnaire):
        if slug in slugs:
            logger.warning("Error to save questionsets and questions of the questionnaire %s ", questionnaire)

            print("Slug: " + slug)
            raise ValueError("Not Unique Slug %s",slug)
        slugs.append(slug)
        return slugs
