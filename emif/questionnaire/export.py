# -*- coding: utf-8 -*-
# Copyright (C) 2014 Universidade de Aveiro, DETI/IEETA, Bioinformatics Group - http://bioinformatics.ua.pt/
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

from questionnaire.models import *

import os
import re

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from questionnaire import QuestionChoices

from utils import TabularProcessor

# TODO: put this code in utils
def removehs(value):
    value = value.replace('h0. ','')
    value = value.replace('h1. ','')
    value = value.replace('h2. ','')
    value = value.replace('h3. ','')
    value = value.replace('h4. ','')
    value = value.replace('h5. ','')
    value = value.replace('h6. ','')
    value = value.replace('h7. ','')
    return value


"""This class is used to export the fingerprint template to excel
"""
class ExportQuestionnaire(object):


    def __init__(self, questionnaire, file_path):
        self.questionnaire = questionnaire
        self.file_path = file_path

    def ignore_questionsets(self, questionset_list):
        raise NotImplementedError("Please Implement this method")

    def export(self):
        raise NotImplementedError("Please Implement this method")


    def level(self, question):
        return question.number.count(".") + 1


    def clean(self, text):
        return removehs(text)


    """This method will build the object according with the type
    of the object to export.
    """
    @staticmethod
    def factory(t_type, questionnaire, file_path):
        if t_type == "csv_plain":
            return ExportQuestionnaireCSVPlain(questionnaire, file_path)

        elif t_type == 'excel':
            return ExportQuestionnaireExcel(questionnaire, file_path)
        else:
            raise Exception("The supplied format is not supported")


class ExportQuestionnaireCSVPlain(ExportQuestionnaire):

    def __init__(self, questionnaire, file_path):
        ExportQuestionnaire.__init__(self, questionnaire)

    def ignore_questionsets(self, questionset_list):
        raise NotImplementedError("Please Implement this method")

    def export(self):
        result = ""
        questionsets = self.questionnaire.questionsets()

        for qs in questionsets:

            result += str(qs.sortid) + " - " + self.clean(qs.text)  + "\n"
            questions = qs.questions()
            for q in questions:
                _level  = self.level(q)
                result += self.get_tabs(_level)
                result += str(q.number)  + " " + self.clean(q.text) +"\n"

        f = None
        if isinstance(file_path, file):
            f = file_path
        else:
            f = open(self.file_path, 'w')

        f.write(result)
        f.close()


class ExportQuestionnaireExcel(ExportQuestionnaire):

    # dependency questions need to be able to translate question numbers into excel line numbers.
    __defaultstyle = NamedStyle(
        name="Default",
        font=Font(name='Verdana', size=8),
        alignment=Alignment(wrap_text=True),
        border=Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
            )
        )

    __bold_default_style = NamedStyle(
        name="DefaultBold",
        font=Font(name='Verdana', size=8, bold=True),
        alignment=Alignment(wrap_text=True),
        border=Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
            )
        )

    __boldstyle = NamedStyle(name="Bold", font=Font(bold=True))

    __headerstyle = NamedStyle(name="Header", alignment=Alignment(horizontal='center'),
            fill=PatternFill(fill_type='solid', start_color='FFCCCCCC'))

    __validatetype = DataValidation(type="list", formula1='"QuestionSet, Category, Question"', allow_blank=True)
    __validateyesno = DataValidation(type="list", formula1='"Yes, No"', allow_blank=True)


    qtype="'Auxiliar Tables'!$E$3:$E$100"

    __validateqtype = DataValidation(type="list", formula1=qtype, allow_blank=True)

    __validatecstate = DataValidation(type="list", formula1='"visible"', allow_blank=True)

    __validatedisp = DataValidation(type="list", formula1='"vertical, horizontal, dropdown"', allow_blank=True)


    def __init__(self, questionnaire, file_path):
        ExportQuestionnaire.__init__(self, questionnaire, file_path)
        self.__ignore_questionsets = []

        self.__number_map = {}

    def __boolean_to_string(self, value):
        if value == True:
            return 'Yes'
        elif value == False:
            return 'No'

        return 'error'

    def __setDefaultStyle(self, _cell):
        _cell.style = self.__defaultstyle

    def __setBold(self, _cell):
        _cell.style = self.__bold_default_style

    def __setHeader(self, _cell):
        _cell.style.font.bold = True

        # Cell background color
        _cell.style = self.__headerstyle

    def __setColumnSizes(self, ws, sizes):
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']

        for i in xrange(len(columns)):
            ws.column_dimensions[columns[i]].width = sizes[i]
    def __getChoices(self, question):
        retstring = ""

        if question.type == 'open-validated':
            try:
                return question.meta()['base']
            except KeyError:
                try:
                    return question.meta()['regex']
                except KeyError:
                    return ""
        elif question.type == 'open-multiple-composition':
            return question.metadata

        if question.type == 'choice-tabular':

            tp = TabularProcessor(question)

            return tp.__unicode__()
        else:
            choices = Choice.objects.filter(question=question)

            for choice in choices:
                extra = ""
                if choice.opt:
                    extra = '{...}'

                retstring+=choice.value+extra+'|'

            if(len(retstring) > 1):
                return retstring[:-1]

        return retstring

    def __getHelp(self, question):
        if question.type == 'open-validated':
            unit = None
            desc = None
            try:
                unit = question.meta()['unit']
            except KeyError:
                unit = ""
            try:
                desc = question.meta()['unit_desc']
            except KeyError:
                desc = ""

            if question.help_text != '':
                return "%s|%s|%s" % (unit, desc, question.help_text)
            else:
                if desc != '':
                    return "%s|%s" % (unit, desc)

                return unit

        return question.help_text

    def __getChoiceNumber(self, parent, option):
        yesno_questions = ['choice-yesno','choice-yesnodontknow']

        if parent.type in yesno_questions:
            if option.lower() == 'yes':
                return "1"
            elif option.lower() == 'no':
                return "2"
            elif option.lower() == 'dontknow':
                return "3"

            return "error"
        else:
            option = option.replace('{...}', '')
            print("Question: " + str(parent))
            print("Question: " + str(parent.pk))
            print("Option: " + str(option))
            return Choice.objects.get(question=parent, value=option).sortid


    def __processDependencies(self, question):

        try:
            if (question.checks == None):
                question.checks = ""
            valid = re.search('.*dependent="([0-9\.]+),(.*)".*', question.checks, re.IGNORECASE)

            if valid:
                number = valid.group(1)
                option = valid.group(2)

                (line, parent) = self.__number_map[number]
                print("DEP: " + str(question.pk))
                optionid = self.__getChoiceNumber(parent, option)

                restring = str(line)+'|'+str(optionid)

                return restring

            # No dependencies
            else:
                return ""
        except KeyError:
            print("-- ERROR: Couldn't find a mapping for question "+str(question.number))
            print("-- ERROR: Couldn't find a mapping for question id "+str(question.pk))
            return "error"

    def __processDisposition(self, disposition):

        if disposition == 0:
            return 'vertical'
        elif disposition == 1:
            return 'horizontal'

        elif disposition == 2:
            return 'dropdown'

        return ''

    def __processCommentVisible(self, visible):

        if visible:
            return 'visible'

        return ''

    def __addQuestion(self, line, ws, question):
        self.__number_map[question.number] = (question.slug, question)

        valid = re.search('(h[0-9])+\. (.*)', question.text_en, re.IGNORECASE)
        choice_types = ['open-validated', 'open-multiple-composition', 'choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform', 'choice-multiple-freeform-options', 'choice-tabular']

        if valid:
            level = str(valid.group(1))
            text = str(valid.group(2))
            type = None

            choices = ''

            if question.category:
                type = 'Category'
            else:
                type = 'Question'

            if question.type in choice_types:
                choices = self.__getChoices(question)

            ws.append([
                    type,
                    text,
                    level,
                    question.type,
                    choices,
                    self.__getHelp(question),
                    self.__boolean_to_string(question.tooltip),
                    question.slug_fk.slug1,
                    self.__processDependencies(question),
                    self.__boolean_to_string(question.stats),
                    self.__processCommentVisible(question.visible_default),
                    self.__processDisposition(question.disposition),
                    self.__boolean_to_string(question.show_advanced)
                ])

            for row in ws.iter_rows(min_col=1, min_row=line, max_col=13, max_row=line):
                for cell in row:
                    self.__setDefaultStyle(cell)

            if question.category:
                self.__setBold(ws.cell(column=2, row=line))

            return True

        return False

    def ignore_questionsets(self, questionset_list):
        self.__ignore_questionsets = questionset_list

    def export(self):
        wb = load_workbook(filename =r'questionnaire/empty2.xlsx')
        ws = wb.get_active_sheet()
        ws.title = "Questionnaire"

        ws.cell(column=2, row=1).value = self.questionnaire.name
        self.__setBold(ws.cell(column=2, row=1))


        # Write questionnaire types to auxiliary excel spreadsheet

        auxiliary = wb['Auxiliar Tables']

        j = 4
        for key, value in QuestionChoices:
            auxiliary.cell(column=5, row=j).value = key
            j+=1


        # for sanity, im keeping a pointer to the row im in...
        pointer = 3

        for questionset in self.questionnaire.questionsets():
            if questionset in self.__ignore_questionsets:
                continue

            help_text = ""
            if questionset.help_text is not None :
                help_text = questionset.help_text.replace('<br />', '\n')

            ws.append(['QuestionSet', questionset.text_en.replace('h1. ',''),
                        questionset.sortid, '', '', help_text,
                        self.__boolean_to_string(questionset.tooltip), questionset.heading,
                        '', '', '','', self.__boolean_to_string(questionset.show_advanced) ])


            for row in ws.iter_rows(min_col=1, min_row=pointer, max_col=11, max_row=pointer):
                for cell in row:
                    self.__setDefaultStyle(cell)

            self.__setBold(ws.cell(column=1, row=pointer))
            self.__setBold(ws.cell(column=2, row=pointer))
            self.__setBold(ws.cell(column=3, row=pointer))

            pointer += 1

            for question in questionset.questions():
                inserted = self.__addQuestion(pointer, ws, question)

                if inserted:
                    pointer += 1
                else:
                    print("-- ERROR PROCESSING QUESTION header for: "+str(question.text_en))
                    print("-- ERROR PROCESSING QUESTION id: "+str(question.pk))
                    break

        # Adding validation data, to create dropdown abilities as the original
        self.__validatetype.ranges.add('A3:A'+str(pointer))
        self.__validateqtype.ranges.add('D3:D'+str(pointer))
        self.__validateyesno.ranges.add('G3:G'+str(pointer))
        self.__validatecstate.ranges.add('K3:K'+str(pointer))
        self.__validatedisp.ranges.add('L3:L'+str(pointer))


        ws.add_data_validation(self.__validatetype)
        ws.add_data_validation(self.__validateyesno)
        ws.add_data_validation(self.__validateqtype)
        ws.add_data_validation(self.__validatecstate)
        ws.add_data_validation(self.__validatedisp)

        # Freezing first two rows
        ws.freeze_panes = ws.cell(column=1, row=3)

        wb.save(self.file_path)



# def main():

#     q = Questionnaire.objects.get(id=53)

#     exporter = ExportQuestionnaire.factory("csv_plain", q, '/tmp/csvplain.csv')
#     exporter.export()









